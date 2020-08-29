import openpyxl as xl


def music_workbook():
    wb = xl.load_workbook('music.xlsx')
    music = wb['music']
    genre = []
    gender = []
    gender_amount = []
    male_genre_counter_save = []
    female_genre_counter_save = []
    for row in range(2, music.max_row + 1):
        genre_cell = music.cell(row, 3)
        gender_cell = music.cell(row, 2)
        gender_amount.append(gender_cell.value)
        if genre_cell.value not in genre:
            genre.append(genre_cell.value)
        if gender_cell.value not in gender:
            gender.append(gender_cell.value)
    for col in range(len(genre)):
        genre_name = music.cell(6, col+6)
        genre_name.value = genre[col]
    for row in range(len(gender)):
        gender_name = music.cell(row + 7, 5)
        gender_name.value = gender[row]

    def counter(number, gender_value):
        for rows in range(2, music.max_row + 1):
            genre_counter_list = music.cell(rows, 3)
            gender_counter_list = music.cell(rows, 2)
            if len(male_genre_counter_save) < gender_amount.count(1) and gender_counter_list.value == 1:
                male_genre_counter_save.append(genre_counter_list.value)
            elif len(female_genre_counter_save) < gender_amount.count(0) and gender_counter_list.value == 0:
                female_genre_counter_save.append(genre_counter_list.value)
        if gender_value == 1:
            return male_genre_counter_save.count(number)
        elif gender_value == 0:
            return female_genre_counter_save.count(number)
    for col in range(len(genre)):
        gender_value = 1
        total_number = counter(genre[col], gender_value)
        genre_counter = music.cell(7, col + 6)
        genre_counter.value = total_number
    for col in range(len(genre)):
        gender_value = 0
        total_number = counter(genre[col], gender_value)
        genre_counter = music.cell(8, col + 6)
        genre_counter.value = total_number

    wb.save('music_ratio.xlsx')
