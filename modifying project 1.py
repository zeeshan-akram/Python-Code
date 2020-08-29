from pathlib import Path
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def workbook_contents(filename, percentage_drop, bar_chart_conformtion):
    wb = xl.load_workbook(file_name)
    sheet = wb["Sheet1"]
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * percentage_drop
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_header = sheet.cell(1, 4)
        corrected_price_header.value = 'corrected price'
        corrected_price_cell.value = corrected_price
    if bar_chart_conformtion == 'yes':
        values = Reference(sheet,
                  min_row=2,
                  max_row=sheet.max_row,
                  min_col=4,
                  max_col=4)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'a5')
        print("Your file saved.")
    else:
        print("ok! Your file have been saved.")

    wb.save('transaction_correct.xlsx')


print("Edit your excel file efficiently...")
file_name = str(input('Enter file name: '))
path = Path(file_name)
if path.exists():
    percentage_drop = int(input("how much percentage do you wanna drop: "))
    percentage_drop /= 100
    bar_chart_conformtion = input("Do you also want bar chart? ").lower()
    workbook_contents(file_name, percentage_drop, bar_chart_conformtion)
else:
    print("Didn't find file")